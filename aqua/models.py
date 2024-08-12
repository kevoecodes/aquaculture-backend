import openpyxl
from django.contrib.auth.models import User
from django.db import models
from django.http import HttpResponse
from openpyxl.utils import get_column_letter


class Device(models.Model):

    device_no = models.CharField(max_length=255)
    created_at = models.DateTimeField(auto_now_add=True)


class Reading(models.Model):

    temperature = models.FloatField(default=0.0)
    ph = models.FloatField(default=0.0)
    turbidity = models.FloatField(default=0.0)
    ammonia = models.FloatField(default=0.0)
    dissolved_oxygen = models.FloatField(default=0.0)
    device = models.ForeignKey(Device, on_delete=models.CASCADE)

    created_at = models.DateTimeField(auto_now_add=True)

    class Meta:
        ordering = ('-created_at',)

    import openpyxl
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    @staticmethod
    def export_data_to_excel(columns, filename='export.xlsx'):
        queryset = Reading.objects.all()
        # Create a workbook and select the active worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Exported Data'

        # Write the header row
        for col_num, column_title in enumerate(columns, 1):
            column_letter = get_column_letter(col_num)
            worksheet[f'{column_letter}1'] = column_title

        # Write data rows
        for row_num, record in enumerate(queryset, 2):
            for col_num, column_title in enumerate(columns, 1):
                column_letter = get_column_letter(col_num)
                worksheet[f'{column_letter}{row_num}'] = getattr(record, column_title)

        # Set the response for the download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'

        # Save the workbook to the response
        workbook.save(response)

        return response


class UserDevice(models.Model):
    user = models.ForeignKey(User, on_delete=models.CASCADE)
    device = models.ForeignKey(Device, on_delete=models.CASCADE)

